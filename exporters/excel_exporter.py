"""
Excel export using openpyxl. Produces a multi-sheet workbook:
data sheet with live SUM formulas, a summary sheet, and an embedded chart image.
"""
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
import logger_setup
from models import ReportMetadata

log = logger_setup.get_logger(__name__)

HEADER_FILL = PatternFill(start_color="4C72B0", end_color="4C72B0", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def build_excel_report(
    output_path: Path,
    metadata: ReportMetadata,
    records,
    chart_paths: list[Path] = None,
) -> Path:
    wb = Workbook()

    # --- Summary sheet ---
    summary_ws = wb.active
    summary_ws.title = "Summary"
    summary_ws["A1"] = metadata.title
    summary_ws["A1"].font = Font(size=16, bold=True)
    summary_ws["A3"] = "Period"
    summary_ws["B3"] = f"{metadata.period_start} to {metadata.period_end}"
    summary_ws["A4"] = "Generated"
    summary_ws["B4"] = metadata.generated_at.strftime("%Y-%m-%d %H:%M")
    summary_ws["A5"] = "Total Records"
    summary_ws["B5"] = metadata.total_records
    summary_ws["A6"] = "Validation Errors"
    summary_ws["B6"] = metadata.validation_error_count
    summary_ws["A7"] = "Validation Warnings"
    summary_ws["B7"] = metadata.validation_warning_count

    if chart_paths:
        row_cursor = 10
        for chart_path in chart_paths:
            img = XLImage(str(chart_path))
            img.width, img.height = 420, 280
            summary_ws.add_image(img, f"A{row_cursor}")
            row_cursor += 18

    # --- Data sheet ---
    data_ws = wb.create_sheet("Data")
    headers = ["Record ID", "Date", "Region", "Category", "Revenue", "Quantity", "Source"]
    for col_idx, header in enumerate(headers, start=1):
        cell = data_ws.cell(row=1, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

    for row_idx, r in enumerate(records, start=2):
        data_ws.cell(row=row_idx, column=1, value=r.record_id)
        data_ws.cell(row=row_idx, column=2, value=str(r.record_date))
        data_ws.cell(row=row_idx, column=3, value=r.region)
        data_ws.cell(row=row_idx, column=4, value=r.category)
        data_ws.cell(row=row_idx, column=5, value=r.revenue)
        data_ws.cell(row=row_idx, column=6, value=r.quantity)
        data_ws.cell(row=row_idx, column=7, value=r.source)

    last_row = len(records) + 1
    total_row = last_row + 2
    data_ws.cell(row=total_row, column=4, value="TOTAL").font = Font(bold=True)
    data_ws.cell(row=total_row, column=5,
                 value=f"=SUM(E2:E{last_row})").font = Font(bold=True)
    data_ws.cell(row=total_row, column=6,
                 value=f"=SUM(F2:F{last_row})").font = Font(bold=True)

    # Auto-width columns
    for col_idx, header in enumerate(headers, start=1):
        data_ws.column_dimensions[get_column_letter(col_idx)].width = max(12, len(header) + 4)

    # Basic conditional formatting: highlight rows with revenue below 100 (potential issue)
    from openpyxl.formatting.rule import CellIsRule
    data_ws.conditional_formatting.add(
        f"E2:E{last_row}",
        CellIsRule(operator="lessThan", formula=["100"],
                   fill=PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"))
    )

    wb.save(output_path)
    log.info(f"Excel report written to {output_path}")
    return output_path